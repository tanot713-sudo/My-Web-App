"""
MA Cost Estimation Generator v5
===============================
เปลี่ยนจาก v4 (สำคัญมาก):
  ★ เขียน "ทุก sheet" จาก input — ไม่มีข้อมูลเก่าของ template ค้างอีกต่อไป
      + Action            (v4 อ่านอย่างเดียว → ข้อมูลโรงพยาบาลค้าง 585 แถว)
      + PM Plan           (v4 ไม่แตะเลย → ค้าง 50 แถว)
      + PM Schdule plan   (v4 ข้าม เพราะ LAYOUT เป็น None → ค้าง)
      + Old/New Abbreviation (v4 ไม่แตะเลย)
      + Outsource         (v4 ไม่แตะเลย)
  ★ LAYOUT["pm"] ถอดโครงสร้างจริงแล้ว:
      col A = Location | เดือน m เริ่มคอลัมน์ 2+(m-1)*12
      ในแต่ละเดือนมี 6 ความถี่ อย่างละ 2 คอลัมน์ (SYS, TT hr):
      D1=+0, W1=+2, M1=+4, M3=+6, M6=+8, Y1=+10
  ★ balance_months() คืน sched[freq][month] แยกความถี่ → ลง PM Schedule ได้ตรงช่อง
  ★ clear_block(reset_style=True) ลบทั้งค่าและเส้นขอบของแถวส่วนเกิน
  ★ โหมด --clean : สร้าง template เปล่า (ฟอร์แมตครบ ไม่มีข้อมูล) ไว้เป็น base ถาวร
  ★ ไม่แก้ฟอร์แมต/merge/สี ของโครงหัวตารางเด็ดขาด

input ต้องมี 6 sheet: PROJECT, EQUIPMENT, PM_PLAN, TOOLS_SPARE, PM_ACTIVITY, OUTSOURCE

วิธีใช้ใน Colab:
    !python generate_v5.py Est_Cost_MA_v5_clean.xlsx input_data_RedLine_Dark.xlsx
    !python generate_v5.py --check input_data_RedLine_Dark.xlsx
    !python generate_v5.py --clean Est_Cost_MA_v3_final.xlsx Est_Cost_MA_v5_clean.xlsx
    !python generate_v5.py --dump  Est_Cost_MA_v3_final.xlsx "PM Schdule plan"
"""

import sys, re, os, shutil
from copy import copy
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Border, PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
import routing as RT


# ══════════════════════════════════════════════════════════════════════════════
#  SECTION 1 — LAYOUT   ★ แก้เลขแถว/คอลัมน์ทุกอย่างที่นี่ที่เดียว ★
# ══════════════════════════════════════════════════════════════════════════════
LAYOUT = {
    "sheets": {
        "summary":  "Summary Cost",
        "tools":    "เครื่องมือช่าง (Tools)",
        "spares":   "อะไหล่และวัสดุสิ้นเปลือง",
        "outsource":"Outsource",
        "tel":      "Telephone fee",
        "mprate":   "Manpower rate",
        "action":   "Action",
        "sum":      "SUM",
        "pm":       "PM Schdule plan",
        "pmplan":   "PM Plan",
        "abbr_old": "Old Abbreviation",
        "abbr_new": "New Abbreviation",
        "roster":   "Roster design",
    },

    "summary": {
        "c_tender": (2, 3), "c_name": (3, 3), "c_value": (4, 3),
        "duration_rows": [14, 17, 18, 22, 24, 28, 30], "duration_col": 4,
        "cell_report": "H24", "cell_travel": "H28", "cell_meeting": "H30",
        "ref_tools_cell":  (12, 9),
        "ref_spares_cell": (13, 9),
        "mp_rows": [
            (33, "Management and Support Group", "management_group", False),
            (34, "Maintenance Project Leader",   "project_leader",   False),
            (35, "Supervisor",                   "supervisor",       True),
            (36, "Engineer",                     "engineer",         True),
            (37, "Technician",                   "technician",       True),
            (38, "Admin",                        "admin",            True),
        ],
        "mp_col_label": 2, "mp_col_count": 3, "mp_col_months": 4,
    },

    "itemlist": {   # ใช้ร่วมกัน Tools + Spares
        "row_start": 3, "title_cell": "A1",
        "c_name": 1, "c_qty": 2, "c_price": 3, "c_amount": 4, "c_system": 5,
        "total_label": "ราคารวม", "total_label_col": 3,
        "cols": (1, 5), "scan_to": 400,
    },

    "outsource": {
        "row_start": 4,
        "c_no": 1, "c_desc": 2, "c_dur": 4, "c_unit1": 5, "c_qty": 6,
        "c_unit2": 7, "c_price": 8, "c_total": 9, "c_remark": 10,
        "cols": (1, 10), "total_cell": (3, 9),   # I3 = SUM(I4:I{last})
    },

    "tel": {
        "rows": [(4, "Maintenance Leader", "project_leader"),
                 (5, "Supervisor",         "supervisor"),
                 (6, "Engineer",           "engineer"),
                 (7, "Technician",         "technician")],
        "c_pos": 2, "c_count": 3, "c_cost": 4,
    },

    "mprate": {
        "rows": [(4, "Management Group Cost",      "management_group", 0),
                 (5, "Maintenance Project Leader", "project_leader",   0),
                 (6, "Supervisor",                 "supervisor",       0),
                 (7, "Engineer",                   "engineer",         2000),
                 (8, "Technician",                 "technician",       2000),
                 (9, "Admin",                      "admin",            0)],
        "c_pos": 2, "c_salary": 3, "c_benefit": 4,
    },

    "action": {
        "row_start": 2, "style_row": 2,
        "c_system": 1, "c_name_th": 2, "c_name_en": 3, "c_code": 4,
        "c_freq": 5, "c_freqcode": 6, "c_activity": 7, "c_worktime": 8,
        "cols": (1, 8),
        # 8 แถวต่อ 1 code — ตามฟอร์แมตเดิมของ template เป๊ะ
        "freqs": ["Daily", "Weekly", "Monthly", "Bi-Monthly",
                  "Quarterly", "Semiannually", "Annually", "Collective"],
    },

    "sum": {
        "row_start": 2, "style_row": 2,
        "c_location": 1, "c_system": 2, "c_name_th": 3, "c_name_en": 4,
        "c_old_code": 5, "c_code": 6,
        "c_qty": 13, "c_workers": 14,
        "freq_cols": {"Daily": 7, "Weekly": 8, "M1": 9, "M3": 10, "M6": 11, "Annually": 12},
        "cols": (1, 14),
    },

    # ─── PM Schdule plan — ถอดโครงสร้างจริงจาก template แล้ว ───────────────────
    #  col A = Location (merge แนวตั้งต่อสถานี)
    #  เดือน m : คอลัมน์เริ่ม = 2 + (m-1)*12
    #  ในเดือนนั้น 6 ความถี่ อย่างละ 2 คอลัมน์ → SYS, TT (hr)
    "pm": {
        "row_start": 5, "style_row": 5,
        "c_location": 1,
        "c_month_1": 2, "month_width": 12,
        "freq_offset": {"Daily": 0, "Weekly": 2, "M1": 4, "M3": 6, "M6": 8, "Annually": 10},
        "cols": (1, 145),
    },

    "pmplan": {
        "row_start": 3,
        "c_item": 1, "c_system": 2, "c_subsystem": 3,
        "freq_cols": {"Daily": 4, "Weekly": 5, "M1": 6, "M3": 7, "M6": 8, "Annually": 9},
        "cols": (1, 9),
    },

    "abbr_old": {
        "row_start": 2,
        "c_system": 1, "c_name_th": 2, "c_name_en": 3, "c_abbr": 4,
        "cols": (1, 4),
    },

    "abbr_new": {
        "row_start": 2, "c_name_en": 1, "c_code": 2, "cols": (1, 2),
    },

    "roster": {
        "row_start": 4,
        "c_no": 1, "c_name": 2, "c_day_1": 3, "n_days": 30,
        "cols": (1, 33),
    },
}

FREQ_SPEC = {
    "Daily":     {"per_month": 30.0,  "per_year": 360,  "fixed": True,  "phases": None},
    "Weekly":    {"per_month": 4.333, "per_year": 52,   "fixed": True,  "phases": None},
    "M1":        {"per_month": 1.0,   "per_year": 12,   "fixed": True,  "phases": None},
    "M3":        {"per_year": 4,  "fixed": False, "phases": [1, 2, 3],            "step": 3},
    "M6":        {"per_year": 2,  "fixed": False, "phases": [1, 2, 3, 4, 5, 6],   "step": 6},
    "Annually":  {"per_year": 1,  "fixed": False, "phases": list(range(1, 13)),   "step": 12},
}

# PM_PLAN / SUM freq  →  Action sheet Frequency label
FREQ_TO_ACTION = {"Daily": "Daily", "Weekly": "Weekly", "M1": "Monthly",
                  "M3": "Quarterly", "M6": "Semiannually", "Annually": "Annually"}

# default กะตามความถี่: ถี่=กลางวัน, ห่าง=กลางคืน (แก้ได้ที่ PM_ACTIVITY คอลัมน์ Shift)
DEFAULT_SHIFT_BY_FREQ = {"Daily": "Day", "Weekly": "Day", "M1": "Day",
                         "M3": "Night", "M6": "Night", "Annually": "Night"}


def resolve_shift(cfg, code, freq, system):
    """ลำดับความสำคัญ: PM_ACTIVITY (เจาะจง code+freq) > SHIFT (ระดับระบบ) > default ตามความถี่"""
    a = cfg["activity"].get(code, {}).get(freq)
    if a and a.get("shift"):
        return a["shift"]
    sysshift = cfg["shift_map"].get(system)
    if sysshift in ("Day", "Night"):
        return sysshift
    return DEFAULT_SHIFT_BY_FREQ.get(freq, "Day")

NO_BORDER = Border()
NO_FILL   = PatternFill()


# ══════════════════════════════════════════════════════════════════════════════
#  SECTION 2 — UTILITIES
# ══════════════════════════════════════════════════════════════════════════════
def w(ws, row, col, val):
    """เขียนเซลล์แบบปลอดภัย: ข้าม MergedCell"""
    c = ws.cell(row=row, column=col)
    if not isinstance(c, MergedCell):
        c.value = val
        return True
    return False


def copy_style(src, dst):
    if src.has_style:
        dst.font = copy(src.font)
        dst.fill = copy(src.fill)
        dst.border = copy(src.border)
        dst.alignment = copy(src.alignment)
        dst.number_format = src.number_format


def clear_block(ws, r0, r1, c0, c1, reset_style=False):
    """ล้างค่าในโซนข้อมูล — reset_style=True จะลบเส้นขอบ/สีด้วย (ใช้กับแถวส่วนเกิน)"""
    if r1 < r0:
        return
    for r in range(r0, r1 + 1):
        for c in range(c0, c1 + 1):
            cell = ws.cell(row=r, column=c)
            if isinstance(cell, MergedCell):
                continue
            cell.value = None
            if reset_style:
                cell.border = NO_BORDER
                cell.fill = NO_FILL


def unmerge_in(ws, r0, r1, c0=None):
    """ถอด merge ที่อยู่ในโซนข้อมูล (กัน merge เก่าทับแถวใหม่)"""
    for mr in list(ws.merged_cells.ranges):
        if mr.min_row >= r0 and mr.max_row <= r1:
            if c0 is None or (mr.min_col == c0 and mr.max_col == c0):
                ws.unmerge_cells(str(mr))


def snapshot_styles(ws, row, c0, c1):
    return {c: copy(ws.cell(row=row, column=c)._style) for c in range(c0, c1 + 1)}


def apply_styles(ws, row, styles):
    for c, st in styles.items():
        cell = ws.cell(row=row, column=c)
        if not isinstance(cell, MergedCell):
            cell._style = copy(st)


# ══════════════════════════════════════════════════════════════════════════════
#  SECTION 3 — LOAD INPUT
# ══════════════════════════════════════════════════════════════════════════════
def load_input(path):
    wb = load_workbook(path, data_only=True)

    # ---- PROJECT ----
    d = {}
    for row in wb["PROJECT"].iter_rows(min_row=3, values_only=True):
        if row[0] and not str(row[0]).startswith("——") and row[2] is not None:
            d[row[0]] = row[2]

    def f(k, dv): return float(d.get(k, dv))
    def i(k, dv): return int(d.get(k, dv))

    project = {
        "tender_no": d.get("tender_no", ""),
        "project_name": str(d.get("project_name", "โครงการ")),
        "project_short_name": str(d.get("project_short_name", "") or ""),
        "project_value_include_vat": f("project_value_include_vat", 0),
        "contract_duration_months": i("contract_duration_months", 12),
        "report_document_cost": f("report_document_cost", 2000),
        "travel_cost": f("travel_cost", 3000),
        "meeting_cost": f("meeting_cost", 15000),
    }
    manpower = {
        "management_group": {"count": i("mgmt_count", 1),        "months": i("mgmt_months", 3),        "salary": f("mgmt_salary", 150000)},
        "project_leader":   {"count": i("leader_count", 1),      "months": i("leader_months", 12),     "salary": f("leader_salary", 50000)},
        "supervisor":       {"count": i("supervisor_count", 2),  "months": i("supervisor_months", 12), "salary": f("supervisor_salary", 30000)},
        "engineer":         {"count": i("engineer_count", 5),    "months": i("engineer_months", 12),   "salary": f("engineer_salary", 24000)},
        "technician":       {"count": i("technician_count", 34), "months": i("technician_months", 12), "salary": f("technician_salary", 19000)},
        "admin":            {"count": i("admin_count", 1),       "months": i("admin_months", 12),      "salary": f("admin_salary", 15000)},
    }
    telephone = {
        "project_leader": {"count": manpower["project_leader"]["count"], "cost": f("tel_leader_cost", 800)},
        "supervisor":     {"count": manpower["supervisor"]["count"],     "cost": f("tel_supervisor_cost", 399)},
        "engineer":       {"count": manpower["engineer"]["count"],       "cost": f("tel_engineer_cost", 399)},
        "technician":     {"count": manpower["technician"]["count"],     "cost": f("tel_technician_cost", 299)},
    }

    # ---- EQUIPMENT ----
    equipment = []
    for row in wb["EQUIPMENT"].iter_rows(min_row=3, values_only=True):
        if not row[0]:
            continue
        equipment.append({
            "location": row[0], "system": row[1],
            "name_th": row[2], "name_en": row[3],
            "code": str(row[4]).strip() if row[4] else "",
            "qty": int(row[5]) if row[5] else 0,
            "workers": int(row[6]) if row[6] else 1,
            "old_code": row[7] or "",
        })

    # ---- PM_PLAN → action_bank[code][freq] = ชั่วโมง/รอบ/หน่วย ----
    action_bank = {}
    freq_col = {2: "Daily", 3: "Weekly", 4: "M1", 5: "M3", 6: "M6", 7: "Annually"}
    for row in wb["PM_PLAN"].iter_rows(min_row=3, values_only=True):
        if not row[0]:
            continue
        bank = {}
        for col, freq in freq_col.items():
            v = row[col] if col < len(row) else None
            try:
                fv = float(v)
                if fv > 0:
                    bank[freq] = fv
            except (TypeError, ValueError):
                pass
        action_bank[str(row[0]).strip()] = bank

    # ---- TOOLS_SPARE ----
    tools, spares = [], []
    for row in wb["TOOLS_SPARE"].iter_rows(min_row=3, values_only=True):
        if row[1]:
            tools.append({"name": row[1], "qty": row[2] or 0, "unit_price": row[3] or 0, "system": row[4] or ""})
        if len(row) > 6 and row[6]:
            spares.append({"name": row[6], "qty": row[7] or 0, "unit_price": row[8] or 0, "system": row[9] or ""})

    # ---- PM_ACTIVITY long format → activity[code][freq] = {"text","shift"} ----
    # freq ในชีตเป็นรหัสภายใน (Daily/Weekly/M1/M3/M6/Annually)
    activity = {}
    if "PM_ACTIVITY" in wb.sheetnames:
        for row in wb["PM_ACTIVITY"].iter_rows(min_row=3, values_only=True):
            if not row[0] or not row[3]:
                continue
            code = str(row[0]).strip()
            freq = str(row[3]).strip()          # คอลัมน์ D = Frequency (รหัส)
            text = row[4] if len(row) > 4 else None
            shift = str(row[5]).strip().upper()[:1] if len(row) > 5 and row[5] else ""
            hr = row[6] if len(row) > 6 else None      # ชม./รอบ สำรอง (งาน CM)
            activity.setdefault(code, {})[freq] = {
                "text": str(text) if text else None,
                "shift": "Night" if shift == "N" else ("Day" if shift == "D" else ""),
                "hr": hr,
            }

    # ---- OUTSOURCE (ใหม่) ----
    outsource = []
    if "OUTSOURCE" in wb.sheetnames:
        for row in wb["OUTSOURCE"].iter_rows(min_row=3, values_only=True):
            if not row[0]:
                continue
            outsource.append({"desc": row[0], "duration": row[1] or 1, "qty": row[2] or 1,
                              "price": row[3], "remark": row[4] or ""})

    # ---- ROUTE / SHIFT / PARAM (ใหม่) ----
    route, shift_map, param = [], {}, dict(RT.DEFAULTS)
    if "ROUTE" in wb.sheetnames:
        for r in wb["ROUTE"].iter_rows(min_row=3, values_only=True):
            if not r[0]:
                continue
            route.append({"name": r[0], "lat": r[1], "lng": r[2],
                          "circuit": r[3], "order": r[4],
                          "km_override": r[5], "min_override": r[6]})
    if "SHIFT" in wb.sheetnames:
        for r in wb["SHIFT"].iter_rows(min_row=3, values_only=True):
            if r[0] and r[1]:
                shift_map[str(r[0]).strip()] = str(r[1]).strip().capitalize()
    if "PARAM" in wb.sheetnames:
        for r in wb["PARAM"].iter_rows(min_row=3, values_only=True):
            if r[0] and r[1] is not None and str(r[0]).strip() in param:
                param[str(r[0]).strip()] = float(r[1])

    return {"project": project, "manpower": manpower, "telephone": telephone,
            "tools": tools, "spare_parts": spares, "outsource": outsource,
            "equipment_list": equipment, "action_bank": action_bank, "activity": activity,
            "route": route, "shift_map": shift_map, "param": param}


# ══════════════════════════════════════════════════════════════════════════════
#  SECTION 4 — CALC (pure functions)
# ══════════════════════════════════════════════════════════════════════════════
def calc_equipment_hours(cfg):
    out = []
    for eq in cfg["equipment_list"]:
        bank = cfg["action_bank"].get(eq["code"], {})
        freq_hours, fixed_monthly, movable, annual = {}, 0.0, [], 0.0
        for freq, spec in FREQ_SPEC.items():
            hpc_unit = bank.get(freq, 0.0)
            if hpc_unit <= 0:
                freq_hours[freq] = 0.0
                continue
            hpc = hpc_unit * eq["qty"]              # ชั่วโมง/รอบ รวมทุกหน่วย
            freq_hours[freq] = hpc
            if spec["fixed"]:
                m = hpc * spec["per_month"]
                fixed_monthly += m
                annual += m * 12
            else:
                movable.append((freq, hpc))
                annual += hpc * spec["per_year"]
        out.append({**eq, "freq_hours": freq_hours,
                    "hpu": bank,                    # ชั่วโมง/รอบ/หน่วย (ใช้ใน Action)
                    "fixed_monthly": round(fixed_monthly, 4),
                    "movable": movable,
                    "annual_hours": round(annual, 4)})
    return out


def balance_months(rows, passes=40):
    """
    กระจาย M3/M6/Annually ลงเดือน 1–12 ให้สม่ำเสมอ (LPT greedy + local search)
    แตกงานระดับ "หน่วยอุปกรณ์" ไม่ใช่ระดับ code
    เพิ่ม key 'sched' = {freq: {month: hours}} ให้แต่ละแถว → ใช้ลง PM Schedule ตรงช่อง
    """
    totals = {m: 0.0 for m in range(1, 13)}

    for r in rows:
        r["sched"] = {f: {} for f in FREQ_SPEC}
        for f in ("Daily", "Weekly", "M1"):
            h = r["freq_hours"].get(f, 0)
            if h > 0:
                for m in range(1, 13):
                    r["sched"][f][m] = h            # ชั่วโมงต่อรอบ — เท่ากันทุกเดือน
        for m in range(1, 13):
            totals[m] += r["fixed_monthly"]

    units = []
    for idx, r in enumerate(rows):
        q = max(r["qty"], 1)
        for freq, hpc_all in r["movable"]:
            h_unit = hpc_all / q
            for _ in range(q):
                units.append({"row": idx, "freq": freq, "h": h_unit,
                              "yr": h_unit * FREQ_SPEC[freq]["per_year"], "m": None})
    units.sort(key=lambda u: -u["yr"])

    sq = lambda t: sum(v * v for v in t.values())

    def best_phase(u, cur):
        spec = FREQ_SPEC[u["freq"]]
        bm, bs = None, None
        for ph in spec["phases"]:
            ms = list(range(ph, 13, spec["step"]))
            trial = dict(cur)
            for m in ms:
                trial[m] += u["h"]
            s = sq(trial)
            if bs is None or s < bs:
                bm, bs = ms, s
        return bm

    for u in units:
        u["m"] = best_phase(u, totals)
        for m in u["m"]:
            totals[m] += u["h"]

    for _ in range(passes):
        changed = False
        for u in units:
            for m in u["m"]:
                totals[m] -= u["h"]
            nm = best_phase(u, totals)
            if nm != u["m"]:
                changed = True
                u["m"] = nm
            for m in u["m"]:
                totals[m] += u["h"]
        if not changed:
            break

    for u in units:
        s = rows[u["row"]]["sched"][u["freq"]]
        for m in u["m"]:
            s[m] = round(s.get(m, 0.0) + u["h"], 3)

    return {m: round(v, 2) for m, v in totals.items()}


def validate(cfg):
    err, warn = [], []
    codes = [e["code"] for e in cfg["equipment_list"]]
    if not codes:
        err.append("EQUIPMENT sheet ว่าง")

    no_plan = sorted({e["code"] for e in cfg["equipment_list"]
                      if not cfg["action_bank"].get(e["code"])})
    if no_plan:
        err.append(f"{len(no_plan)} code ไม่มีใน PM_PLAN → ชั่วโมงเป็น 0: {no_plan[:8]}")

    zero_qty = [e["code"] for e in cfg["equipment_list"] if e["qty"] <= 0]
    if zero_qty:
        warn.append(f"{len(zero_qty)} รายการมี qty = 0: {zero_qty[:8]}")

    orphan = [c for c in cfg["action_bank"] if c not in codes]
    if orphan:
        warn.append(f"{len(orphan)} code อยู่ใน PM_PLAN แต่ไม่มีใน EQUIPMENT: {orphan[:8]}")

    if not cfg["activity"]:
        warn.append("ไม่มี sheet PM_ACTIVITY → คอลัมน์ PM Activity ใน Action จะว่าง")
    else:
        no_act = sorted({c for c in set(codes) if not cfg["activity"].get(c)})
        if no_act:
            warn.append(f"{len(no_act)} code ไม่มีข้อความ PM Activity: {no_act[:8]}")

    if not cfg["tools"]:
        warn.append("TOOLS_SPARE: ไม่มี tools")
    if not cfg["spare_parts"]:
        warn.append("TOOLS_SPARE: ไม่มี spare parts")
    if not cfg["outsource"]:
        warn.append("ไม่มี sheet OUTSOURCE → sheet Outsource จะว่าง")
    return err, warn


# ══════════════════════════════════════════════════════════════════════════════
#  SECTION 5 — PATCH
# ══════════════════════════════════════════════════════════════════════════════
def patch_summary(wb, cfg):
    L = LAYOUT["summary"]
    ws = wb[LAYOUT["sheets"]["summary"]]
    p, mp = cfg["project"], cfg["manpower"]
    mo = p["contract_duration_months"]

    w(ws, *L["c_tender"], p["tender_no"])
    w(ws, *L["c_name"],   p["project_name"])
    w(ws, *L["c_value"],  p["project_value_include_vat"])
    for r in L["duration_rows"]:
        w(ws, r, L["duration_col"], mo)
    ws[L["cell_report"]]  = p["report_document_cost"]
    ws[L["cell_travel"]]  = p["travel_cost"]
    ws[L["cell_meeting"]] = p["meeting_cost"]

    for row, label, key, show_n in L["mp_rows"]:
        cnt, dur = mp[key]["count"], mp[key]["months"]
        w(ws, row, L["mp_col_label"],  f"{label} x{cnt}" if show_n else label)
        w(ws, row, L["mp_col_count"],  cnt)
        w(ws, row, L["mp_col_months"], dur)
    print("  ✅ Summary Cost")


def _patch_itemlist(wb, sheet_key, items, ref_cell, title):
    L = LAYOUT["itemlist"]
    sheet = LAYOUT["sheets"][sheet_key]
    ws = wb[sheet]
    start, (c0, c1) = L["row_start"], L["cols"]

    w(ws, 1, 1, title)                                  # ล้างหัวเรื่องเก่า (โรงพยาบาล)
    old_last = max(ws.max_row, start)
    styles = snapshot_styles(ws, start, c0, c1)
    clear_block(ws, start, old_last, c0, c1)            # ล้างค่าเก่าทั้งหมด รวม "ราคารวม" ค้าง

    for n, it in enumerate(items):
        r = start + n
        apply_styles(ws, r, styles)
        w(ws, r, L["c_name"],   it["name"])
        w(ws, r, L["c_qty"],    it["qty"])
        w(ws, r, L["c_price"],  it["unit_price"])
        w(ws, r, L["c_amount"], f"=B{r}*C{r}")
        w(ws, r, L["c_system"], it.get("system", ""))

    total_row = start + len(items)
    apply_styles(ws, total_row, styles)
    w(ws, total_row, L["total_label_col"], L["total_label"])
    w(ws, total_row, L["c_amount"], f"=SUM(D{start}:D{total_row - 1})")

    clear_block(ws, total_row + 1, old_last, c0, c1, reset_style=True)
    w(wb[LAYOUT["sheets"]["summary"]], *ref_cell, f"='{sheet}'!D{total_row}")
    print(f"  ✅ {sheet}  ({len(items)} รายการ, total row {total_row})")


def patch_tools(wb, cfg):
    _patch_itemlist(wb, "tools", cfg["tools"], LAYOUT["summary"]["ref_tools_cell"],
                    "รายการเครื่องมือช่าง (Tools)")


def patch_spares(wb, cfg):
    _patch_itemlist(wb, "spares", cfg["spare_parts"], LAYOUT["summary"]["ref_spares_cell"],
                    "รายการอะไหล่และวัสดุสิ้นเปลือง (Spare Parts)")


def patch_outsource(wb, cfg):
    L = LAYOUT["outsource"]
    ws = wb[LAYOUT["sheets"]["outsource"]]
    start, (c0, c1) = L["row_start"], L["cols"]
    old_last = max(ws.max_row, start)
    styles = snapshot_styles(ws, start, c0, c1)
    clear_block(ws, start, old_last, c0, c1)

    items = cfg["outsource"]
    for n, it in enumerate(items):
        r = start + n
        apply_styles(ws, r, styles)
        w(ws, r, L["c_no"],     n + 1)
        w(ws, r, L["c_desc"],   it["desc"])
        w(ws, r, L["c_dur"],    it["duration"])
        w(ws, r, L["c_unit1"],  "Time")
        w(ws, r, L["c_qty"],    it["qty"])
        w(ws, r, L["c_unit2"],  "Time")
        w(ws, r, L["c_price"],  it["price"])
        w(ws, r, L["c_total"],  f"=H{r}*F{r}*D{r}")
        w(ws, r, L["c_remark"], it["remark"])

    last = start + len(items) - 1
    w(ws, *L["total_cell"], f"=SUM(I{start}:I{max(last, start)})")
    clear_block(ws, last + 1, old_last, c0, c1, reset_style=True)
    print(f"  ✅ Outsource  ({len(items)} รายการ)")


def patch_telephone(wb, cfg):
    L = LAYOUT["tel"]
    ws = wb[LAYOUT["sheets"]["tel"]]
    for row, pos, key in L["rows"]:
        w(ws, row, L["c_pos"],   pos)
        w(ws, row, L["c_count"], cfg["telephone"][key]["count"])
        w(ws, row, L["c_cost"],  cfg["telephone"][key]["cost"])
    print("  ✅ Telephone fee")


def patch_mprate(wb, cfg):
    L = LAYOUT["mprate"]
    ws = wb[LAYOUT["sheets"]["mprate"]]
    for row, pos, key, benefit in L["rows"]:
        w(ws, row, L["c_pos"],     pos)
        w(ws, row, L["c_salary"],  cfg["manpower"][key]["salary"] - benefit)
        w(ws, row, L["c_benefit"], benefit or None)
    print("  ✅ Manpower rate")


# ---- Action ------------------------------------------------------------------
ACTION_FREQ_SRC = {"Daily": "Daily", "Weekly": "Weekly", "Monthly": "M1",
                   "Quarterly": "M3", "Semiannually": "M6", "Annually": "Annually"}


def patch_action(wb, cfg, rows):
    """
    สร้าง Action sheet ใหม่ทั้งแผ่นจาก input — 8 แถว/code ตามฟอร์แมตเดิม
    คืน action_map[code][freq_label] = {'row': n}
    """
    L = LAYOUT["action"]
    ws = wb[LAYOUT["sheets"]["action"]]
    start, (c0, c1) = L["row_start"], L["cols"]
    old_last = max(ws.max_row, start)
    styles = snapshot_styles(ws, L["style_row"], c0, c1)
    clear_block(ws, start, old_last, c0, c1)

    # code ที่ไม่ซ้ำ เรียงตามลำดับที่พบใน EQUIPMENT (จัดกลุ่มตามระบบ)
    seen, uniq = set(), []
    for eq in rows:
        if eq["code"] not in seen:
            seen.add(eq["code"])
            uniq.append(eq)

    action_map, r, last_sys = {}, start, None
    for eq in uniq:
        code = eq["code"]
        acts = cfg["activity"].get(code, {})
        for n, freq in enumerate(L["freqs"]):
            apply_styles(ws, r, styles)
            if n == 0:
                if eq["system"] != last_sys:
                    w(ws, r, L["c_system"], eq["system"])
                    last_sys = eq["system"]
                w(ws, r, L["c_name_th"], eq["name_th"])
                w(ws, r, L["c_name_en"], eq["name_en"])
            w(ws, r, L["c_code"],     code)
            w(ws, r, L["c_freq"],     freq)
            w(ws, r, L["c_freqcode"], f'=D{r}&"-"&E{r}')
            src = ACTION_FREQ_SRC.get(freq, freq)           # freq(Action) → รหัสภายใน
            act = acts.get(src, {})
            w(ws, r, L["c_activity"], act.get("text"))
            hpu = eq["hpu"].get(src)                        # ชม.จาก PM_PLAN
            if hpu is None:
                hpu = act.get("hr")                        # สำรอง: งาน CM จาก PM_ACTIVITY
            w(ws, r, L["c_worktime"], hpu)
            ws.row_dimensions[r].height = None      # ปล่อยให้ Excel auto-fit ตามข้อความ
            action_map.setdefault(code, {})[freq] = {"row": r}
            r += 1

    clear_block(ws, r, old_last, c0, c1, reset_style=True)
    print(f"  ✅ Action  ({len(uniq)} code x 8 ความถี่ = {r - start} แถว)")
    return action_map


# ---- SUM ---------------------------------------------------------------------
def patch_sum(wb, cfg, rows, action_map):
    L = LAYOUT["sum"]
    ws = wb[LAYOUT["sheets"]["sum"]]
    start, (c0, c1) = L["row_start"], L["cols"]
    old_last = max(ws.max_row, start)
    styles = snapshot_styles(ws, L["style_row"], c0, c1)
    clear_block(ws, start, old_last, c0, c1)

    for n, eq in enumerate(rows):
        r = start + n
        apply_styles(ws, r, styles)
        w(ws, r, L["c_location"], eq["location"])
        w(ws, r, L["c_system"],   eq["system"])
        w(ws, r, L["c_name_th"],  eq["name_th"])
        w(ws, r, L["c_name_en"],  eq["name_en"])
        w(ws, r, L["c_old_code"], eq["old_code"] or None)
        w(ws, r, L["c_code"],     eq["code"])
        w(ws, r, L["c_qty"],      eq["qty"])
        w(ws, r, L["c_workers"],  eq["workers"])

        amap = action_map.get(eq["code"], {})
        for freq, col in L["freq_cols"].items():
            info = amap.get(FREQ_TO_ACTION[freq])
            if info and eq["hpu"].get(freq):
                w(ws, r, col, f"=Action!H{info['row']}*M{r}")   # ผูกสูตรกับ Action
            else:
                w(ws, r, col, None)

    last = start + len(rows) - 1
    clear_block(ws, last + 1, old_last, c0, c1, reset_style=True)
    print(f"  ✅ SUM  ({len(rows)} แถว)")
    return {n: start + n for n in range(len(rows))}   # index → sum row


# ---- PM Schedule -------------------------------------------------------------
def patch_pm_schedule(wb, cfg, rows, sum_rowmap, month_totals):
    L = LAYOUT["pm"]
    ws = wb[LAYOUT["sheets"]["pm"]]
    start, (c0, c1) = L["row_start"], L["cols"]
    old_last = max(ws.max_row, start)

    unmerge_in(ws, start, old_last, c0=L["c_location"])
    styles = snapshot_styles(ws, L["style_row"], c0, min(c1, ws.max_column))
    clear_block(ws, start, old_last, c0, min(c1, ws.max_column))

    # เรียงตามลำดับเดิมของ EQUIPMENT → แถว PM ตรงกับแถว SUM แบบ 1:1
    r = start
    loc_groups, cur_loc, loc_start = [], None, start
    for idx, eq in enumerate(rows):
        if eq["location"] != cur_loc:
            if cur_loc is not None:
                loc_groups.append((cur_loc, loc_start, r - 1))
            cur_loc, loc_start = eq["location"], r

        apply_styles(ws, r, styles)
        srow = sum_rowmap[idx]
        for freq, off in L["freq_offset"].items():
            sched = eq["sched"].get(freq, {})
            for m, hrs in sched.items():
                if not hrs:
                    continue
                base = L["c_month_1"] + (m - 1) * L["month_width"] + off
                w(ws, r, base,     f"=SUM!F{srow}")     # SYS = รหัสอุปกรณ์ (ผูกกับ SUM)
                w(ws, r, base + 1, round(hrs, 2))       # TT (hr) ของเดือนนั้น
        r += 1
    if cur_loc is not None:
        loc_groups.append((cur_loc, loc_start, r - 1))

    for loc, r0, r1 in loc_groups:
        if r1 > r0:
            ws.merge_cells(start_row=r0, start_column=1, end_row=r1, end_column=1)
        ws.cell(row=r0, column=1).value = loc

    clear_block(ws, r, old_last, c0, min(c1, ws.max_column), reset_style=True)

    lo, hi = min(month_totals.values()), max(month_totals.values())
    avg = sum(month_totals.values()) / 12 or 1
    print(f"  ✅ PM Schdule plan  ({len(rows)} แถว, {len(loc_groups)} สถานที่)")
    print(f"      โหลด/เดือน min={lo:,.0f} max={hi:,.0f} avg={avg:,.0f} → กระจาย ±{(hi-lo)/avg*100:.1f}%")


# ---- PM Plan -----------------------------------------------------------------
def patch_pm_plan(wb, cfg, rows):
    L = LAYOUT["pmplan"]
    ws = wb[LAYOUT["sheets"]["pmplan"]]
    start, (c0, c1) = L["row_start"], L["cols"]
    old_last = max(ws.max_row, start)
    styles = snapshot_styles(ws, start, c0, c1)
    clear_block(ws, start, old_last, c0, c1)

    # จัดกลุ่ม: system → [(name_th, hpu)] ไม่ซ้ำ
    systems, seen = {}, set()
    for eq in rows:
        if eq["code"] in seen:
            continue
        seen.add(eq["code"])
        systems.setdefault(eq["system"], []).append(eq)

    r, item = start, 1
    for sys_name, eqs in systems.items():
        apply_styles(ws, r, styles)
        w(ws, r, L["c_item"], item)
        w(ws, r, L["c_system"], sys_name)
        for freq, col in L["freq_cols"].items():       # ระบบ = union ของอุปกรณ์ใต้ระบบ
            if any(e["hpu"].get(freq) for e in eqs):
                w(ws, r, col, "X")
        r += 1
        item += 1
        for e in eqs:
            apply_styles(ws, r, styles)
            w(ws, r, L["c_subsystem"], e["name_th"])
            for freq, col in L["freq_cols"].items():
                if e["hpu"].get(freq):
                    w(ws, r, col, "X")
            r += 1

    clear_block(ws, r, old_last, c0, c1, reset_style=True)
    print(f"  ✅ PM Plan  ({len(systems)} ระบบ, {r - start} แถว)")


# ---- Abbreviation ------------------------------------------------------------
def patch_abbreviation(wb, cfg, rows):
    Lo = LAYOUT["abbr_old"]
    ws = wb[LAYOUT["sheets"]["abbr_old"]]
    start, (c0, c1) = Lo["row_start"], Lo["cols"]
    old_last = max(ws.max_row, start)
    unmerge_in(ws, start, old_last)
    styles = snapshot_styles(ws, start, c0, c1)
    clear_block(ws, start, old_last, c0, c1)

    seen, uniq = set(), []
    for eq in rows:
        if eq["code"] not in seen:
            seen.add(eq["code"])
            uniq.append(eq)

    r, last_sys = start, None
    for eq in uniq:
        apply_styles(ws, r, styles)
        if eq["system"] != last_sys:
            w(ws, r, Lo["c_system"], eq["system"])
            last_sys = eq["system"]
        w(ws, r, Lo["c_name_th"], eq["name_th"])
        w(ws, r, Lo["c_name_en"], eq["name_en"])
        w(ws, r, Lo["c_abbr"],    eq["old_code"] or eq["code"])
        r += 1
    clear_block(ws, r, old_last, c0, c1, reset_style=True)

    Ln = LAYOUT["abbr_new"]
    ws2 = wb[LAYOUT["sheets"]["abbr_new"]]
    s2, (d0, d1) = Ln["row_start"], Ln["cols"]
    old2 = max(ws2.max_row, s2)
    st2 = snapshot_styles(ws2, s2, d0, d1)
    clear_block(ws2, s2, old2, d0, d1)
    r2 = s2
    for eq in uniq:
        apply_styles(ws2, r2, st2)
        w(ws2, r2, Ln["c_name_en"], eq["name_en"])
        w(ws2, r2, Ln["c_code"],    eq["code"])
        r2 += 1
    clear_block(ws2, r2, old2, d0, d1, reset_style=True)
    print(f"  ✅ Old/New Abbreviation  ({len(uniq)} code)")


# ---- Roster ------------------------------------------------------------------
def patch_roster(wb, cfg):
    L = LAYOUT["roster"]
    ws = wb[LAYOUT["sheets"]["roster"]]
    mp = cfg["manpower"]
    start, (c0, c1) = L["row_start"], L["cols"]
    old_last = max(ws.max_row, start)

    staff = (["Maintenance Leader"]
             + [f"Supervisor Staff {i+1}" for i in range(mp["supervisor"]["count"])]
             + ["Admin"]
             + [f"Engineer Staff {i+1}"   for i in range(mp["engineer"]["count"])]
             + [f"Technical Staff {i+1}"  for i in range(mp["technician"]["count"])])

    styles = snapshot_styles(ws, start, c0, c1)
    clear_block(ws, start, old_last, c0, c1)

    pattern = ["M", "M", "M", "M", "M", "X", "X"]   # TODO: ใส่กะจริงเมื่อยืนยัน shift
    for n, name in enumerate(staff):
        r = start + n
        apply_styles(ws, r, styles)
        w(ws, r, L["c_no"],   n + 1)
        w(ws, r, L["c_name"], name)
        for d in range(L["n_days"]):
            w(ws, r, L["c_day_1"] + d, pattern[(d + n) % 7])

    clear_block(ws, start + len(staff), old_last, c0, c1, reset_style=True)
    print(f"  ✅ Roster design  ({len(staff)} คน x {L['n_days']} วัน)")



# ---- PM Shift Plan (ชีตใหม่) --------------------------------------------------
def patch_shift_plan(wb, cfg, rows):
    """ชีต 'PM Shift Plan' — แบ่งกะเช้า/บ่าย/ดึก + กลุ่มสถานี + จำนวนคนที่ต้องใช้จริง"""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    thin = Border(*[Side(style="thin")] * 4)
    HF = PatternFill("solid", start_color="BDD7EE", end_color="BDD7EE")
    TF = PatternFill("solid", start_color="1F3864", end_color="1F3864")
    DF = PatternFill("solid", start_color="FCE4D6", end_color="FCE4D6")  # กลางวัน
    NF = PatternFill("solid", start_color="D9D2E9", end_color="D9D2E9")  # กลางคืน

    if "PM Shift Plan" in wb.sheetnames:
        del wb["PM Shift Plan"]
    ws = wb.create_sheet("PM Shift Plan")
    p = cfg["param"]

    # รวมชั่วโมง/ปี แยกสถานี × กะ (ตาม activity shift)
    from collections import defaultdict
    day = defaultdict(float)
    night = defaultdict(float)
    for eq in rows:
        loc = str(eq["location"])
        for freq, sched in eq["sched"].items():
            hy = sum(sched.values()) * RT.ROUNDS_PER_MONTH.get(freq, 1.0)  # ชม./ปี
            if hy <= 0:
                continue
            sh = RT._shift_of(eq, freq, cfg)
            (night if sh == "Night" else day)[loc] += hy

    def title(r, text, span=8):
        c = ws.cell(row=r, column=1, value=text)
        c.font = Font(bold=True, color="FFFFFF", size=11); c.fill = TF
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=span)

    def header(r, cols):
        for i, h in enumerate(cols, 1):
            c = ws.cell(row=r, column=i, value=h)
            c.font = Font(bold=True, size=10); c.fill = HF
            c.border = thin; c.alignment = Alignment(horizontal="center", wrap_text=True)

    wd = p["workdays_month"]
    dsh, nsh = p["day_shift_hours"], p["night_shift_hours"]
    crew = int(p["crew_size"])
    relief_day, relief_night = 1.4, 2.2

    # A. นิยามกะ
    title(1, "A. โครงสร้างกะทำงาน")
    header(2, ["กะ", "ช่วงเวลา", "ประเภทงาน", "ชม./กะ", "relief factor", "", "", ""])
    shifts = [("กะเช้า (A)", "08:00–16:00", "งานในสถานี — ไม่ต้องปิดบริการ", dsh, relief_day),
              ("กะบ่าย (B)", "13:00–21:00", "งานในสถานี — ไม่ต้องปิดบริการ", dsh, relief_day),
              ("กะดึก (C)", "00:30–04:30", "งานลงราง/ปิดบริการ — engineering hours", nsh, relief_night)]
    r = 3
    for s in shifts:
        for i, v in enumerate(s, 1):
            c = ws.cell(row=r, column=i, value=v); c.border = thin; c.font = Font(size=9)
            if i == 1:
                c.fill = NF if "ดึก" in s[0] else DF
        r += 1

    # B. โหลดงานต่อสถานี แยกกะ
    r += 1
    title(r, "B. ปริมาณงานต่อสถานี แยกกะ (ชม./เดือน)")
    r += 1
    header(r, ["สถานที่", "งานกลางวัน (ชม./ด.)", "งานกลางคืน (ชม./ด.)",
               "วัน-ทีมกลางวัน/ด.", "วัน-ทีมกลางคืน/ด.", "", "", ""])
    r += 1
    locs = sorted(set(list(day) + list(night)))
    tot_d = tot_n = 0.0
    for loc in locs:
        dm, nm = day[loc] / 12, night[loc] / 12
        tot_d += dm; tot_n += nm
        vals = [loc, round(dm, 1), round(nm, 1),
                round(dm / dsh, 2), round(nm / nsh, 2)]
        for i, v in enumerate(vals, 1):
            c = ws.cell(row=r, column=i, value=v); c.border = thin; c.font = Font(size=9)
        r += 1
    for i, v in enumerate(["รวม", round(tot_d, 1), round(tot_n, 1),
                           round(tot_d / dsh, 2), round(tot_n / nsh, 2)], 1):
        c = ws.cell(row=r, column=i, value=v); c.border = thin
        c.font = Font(size=9, bold=True); c.fill = HF
    r += 2

    # C. สรุปกำลังคนที่ต้องใช้
    title(r, "C. กำลังคนขั้นต่ำ (รวม relief factor)")
    r += 1
    header(r, ["กะ", "วัน-ทีม/เดือน", "ทีมพร้อมกัน", "ช่าง/ทีม", "ตำแหน่งประจำ",
               "relief factor", "ช่างจริงที่ต้องจ้าง", ""])
    r += 1
    team_day = (tot_d / dsh) / wd
    team_night = (tot_n / nsh) / 30          # กลางคืนทำ 30 วัน (รถไฟวิ่งทุกวัน)
    lines = [("กะเช้า+บ่าย (A+B)", tot_d / dsh, team_day, crew, relief_day),
             ("กะดึก (C)", tot_n / nsh, team_night, crew, relief_night)]
    grand = 0
    for name, vt, teams, cr, rel in lines:
        posts = teams * cr
        real = posts * rel
        grand += real
        vals = [name, round(vt, 1), round(teams, 2), cr, round(posts, 1),
                rel, round(real, 1)]
        for i, v in enumerate(vals, 1):
            c = ws.cell(row=r, column=i, value=v); c.border = thin; c.font = Font(size=9)
        r += 1
    c = ws.cell(row=r, column=1, value="รวมช่างที่ต้องจ้าง (Technician)")
    c.font = Font(size=10, bold=True); c.fill = HF
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    c2 = ws.cell(row=r, column=7, value=round(grand, 1))
    c2.font = Font(size=10, bold=True); c2.fill = HF; c2.border = thin

    for col, w_ in zip("ABCDEFGH", [30, 20, 18, 12, 16, 14, 20, 6]):
        ws.column_dimensions[col].width = w_
    ws.freeze_panes = "A2"

    print(f"  ✅ PM Shift Plan  (กลางวัน {tot_d:.0f} ชม./ด. + กลางคืน {tot_n:.0f} ชม./ด. "
          f"→ ช่าง ~{grand:.0f} คน รวม relief)")


# ---- Routing (ชีตภาพเส้นทาง) --------------------------------------------------
def patch_routing_images(wb, cfg, out_dir):
    """
    ชีต 'Routing' — ภาพเส้นทางฝังในไฟล์ Excel เลย พร้อมหัวข้อและระยะทางรวม
    ฟอร์แมตตามเอกสารอ้างอิงของ AMR:
        PM Routing no.1        ระยะทางรวม : 209 km.
                               เวลารวม : 176 min or 2.93 hr.
        [ภาพเส้นทาง]
    """
    if not cfg["route"]:
        return
    from openpyxl.styles import Font, Alignment
    try:
        from openpyxl.drawing.image import Image as XLImage
    except ImportError:
        print("  ⏭️  Routing — ข้าม (openpyxl ไม่มีโมดูลรูปภาพ)")
        return

    p = cfg["param"]
    circuits = RT.build_circuits(cfg["route"], p)
    if not circuits:
        return

    if "Routing" in wb.sheetnames:
        del wb["Routing"]
    ws = wb.create_sheet("Routing")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 3
    for col in "BCDE":
        ws.column_dimensions[col].width = 14

    IMG_H = 300                      # ความสูงภาพในชีต — ล็อกความสูงให้ทุกวงเท่ากัน
    IMG_W_MAX = 1450                 # ถ้ากว้างเกินนี้ ย่อลงทั้งภาพ
    ROW_PX = 20                      # ความสูงแถวมาตรฐาน
    svg_dir = os.path.join(out_dir, "route_svg")

    r = 3
    ok = 0
    for i, (cid, c) in enumerate(circuits.items(), 1):
        # ── หัวข้อ ──
        h = ws.cell(row=r, column=2, value=f"PM Routing no.{i}")
        h.font = Font(size=18, bold=True, underline="single", color="1F3864")

        # ── สรุประยะทาง/เวลา ──
        hrs = c["min_day"] / 60
        s1 = ws.cell(row=r, column=6, value=f'ระยะทางรวม : {c["km"]:.1f} km.')
        s2 = ws.cell(row=r + 1, column=6, value=f'เวลารวม : {c["min_day"]:.0f} min or {hrs:.2f} hr.')
        for s in (s1, s2):
            s.font = Font(size=11, bold=True, color="1F3864")

        # ── ชื่อวงเต็ม (บอกว่าครอบคลุมสถานีไหน) ──
        n = ws.cell(row=r + 2, column=2, value=cid + "  •  " + " → ".join(c["stops"]))
        n.font = Font(size=10, italic=True, color="595959")

        # ── ภาพ ──
        svg = RT.render_svg(cid, c, p, svg_dir, caption=False)
        png = RT.render_png(svg, scale=1.6)
        if png:
            img = XLImage(png)
            w0, h0 = img.width, img.height
            h = IMG_H
            w = h * w0 / h0
            if w > IMG_W_MAX:                       # ภาพกว้างมาก (สถานีเยอะ) → ย่อทั้งภาพ
                w, h = IMG_W_MAX, IMG_W_MAX * h0 / w0
            img.width, img.height = int(w), int(h)
            img.anchor = f"B{r + 4}"
            ws.add_image(img)
            rows_used = int(h / ROW_PX) + 3
            ok += 1
        else:
            ws.cell(row=r + 4, column=2,
                    value="(ไม่สามารถแปลงภาพได้ — ดูไฟล์ .svg ในโฟลเดอร์ route_svg)")
            rows_used = 4
        r += 4 + rows_used

    print(f"  ✅ Routing  ({ok}/{len(circuits)} ภาพฝังในไฟล์)")


# ---- PM Routing (ชีตใหม่) -----------------------------------------------------
def patch_pm_routing(wb, cfg, rows, svg_dir=None):
    """ชีต 'PM Routing' — โมเดลเส้นทางมาตรฐาน + export ภาพ SVG ทุกวง"""
    if not cfg["route"]:
        print("  ⏭️  PM Routing — ข้าม (ไม่มีชีต ROUTE)")
        return

    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    thin = Border(*[Side(style="thin")] * 4)
    HF = PatternFill("solid", start_color="BDD7EE", end_color="BDD7EE")
    TF = PatternFill("solid", start_color="1F3864", end_color="1F3864")
    NF = PatternFill("solid", start_color="FFF2CC", end_color="FFF2CC")

    if "PM Routing" in wb.sheetnames:
        del wb["PM Routing"]
    ws = wb.create_sheet("PM Routing")

    p = cfg["param"]
    circuits = RT.build_circuits(cfg["route"], p)
    plan, summ = RT.plan_year(circuits, rows, cfg["shift_map"], p, cfg)

    def title(r, text, span=9):
        c = ws.cell(row=r, column=1, value=text)
        c.font = Font(bold=True, color="FFFFFF", size=11); c.fill = TF
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=span)

    def header(r, cols):
        for i, h in enumerate(cols, 1):
            c = ws.cell(row=r, column=i, value=h)
            c.font = Font(bold=True, size=10); c.fill = HF
            c.border = thin; c.alignment = Alignment(horizontal="center", wrap_text=True)

    def row(r, vals, fill=None):
        for i, v in enumerate(vals, 1):
            c = ws.cell(row=r, column=i, value=v)
            c.border = thin; c.font = Font(size=9)
            if fill:
                c.fill = fill

    # A. พารามิเตอร์
    title(1, "A. พารามิเตอร์ (แก้ที่ชีต PARAM ของ input)")
    r = 2
    for k, v in p.items():
        ws.cell(row=r, column=1, value=k).font = Font(size=9)
        ws.cell(row=r, column=2, value=v).font = Font(size=9, bold=True)
        r += 1

    # B. เส้นทางมาตรฐาน + ระยะแต่ละขา
    r += 1
    title(r, f"B. เส้นทางมาตรฐาน — ออกจาก {RT.HQ_NAME} ({RT.HQ_LAT}, {RT.HQ_LNG})")
    r += 1
    header(r, ["วง (Circuit)", "ขาที่", "จาก", "ถึง", "ระยะทาง (km)",
               "เวลา กลางวัน (นาที)", "เวลา กลางคืน (นาที)", "", ""])
    r += 1
    for cid, c in circuits.items():
        for i, (a, b, km, md, mn) in enumerate(c["legs"], 1):
            row(r, [cid if i == 1 else "", i, a, b, round(km, 1), round(md), round(mn)])
            r += 1
        row(r, [f"รวม {cid}", "", "", "", round(c["km"], 1),
                round(c["min_day"]), round(c["min_night"])], HF)
        for i in range(1, 8):
            ws.cell(row=r, column=i).font = Font(size=9, bold=True)
        r += 1

    # C. จำนวนรอบรายเดือน
    r += 1
    title(r, "C. จำนวนรอบต่อเดือน (เหลือง = กะกลางคืน / engineering hours)")
    r += 1
    header(r, ["เดือน", "วง (Circuit)", "กะ", "ชม.งานที่ครบกำหนด", "ชม.เดินทาง/รอบ",
               "ชม.ทำงานได้/รอบ", "จำนวนรอบ", "ระยะทางรวม (km)", "ชม.เดินทางรวม"])
    r += 1
    for m in range(1, 13):
        for e in plan[m]:
            row(r, [m, e["circuit"], e["shift"], e["work_h"], e["travel_h"],
                    e["capacity"], e["rounds"], e["km"], e["travel_total"]],
                NF if e["shift"] == "Night" else None)
            r += 1
    row(r, ["รวมทั้งปี", "", "", round(summ["work_h"]), "", "", summ["rounds"],
            round(summ["km"]), round(summ["travel_h"])], HF)
    for i in range(1, 10):
        ws.cell(row=r, column=i).font = Font(size=9, bold=True)

    for col, wd in zip("ABCDEFGHI", [10, 28, 9, 18, 15, 16, 11, 15, 15]):
        ws.column_dimensions[col].width = wd
    ws.freeze_panes = "A2"

    # export ภาพเส้นทาง
    paths = []
    if svg_dir:
        for cid, c in circuits.items():
            paths.append(RT.render_svg(cid, c, p, svg_dir))

    pct = summ["travel_h"] / (summ["travel_h"] + summ["work_h"]) * 100
    print(f"  ✅ PM Routing  ({len(circuits)} วง, {summ['rounds']:,} รอบ/ปี, "
          f"{summ['km']:,.0f} km, เดินทาง {summ['travel_h']:,.0f} ชม. = {pct:.0f}%)")
    for pth in paths:
        print(f"      🖼️  {os.path.basename(pth)}")


# ══════════════════════════════════════════════════════════════════════════════
#  SECTION 6 — CLEAN TEMPLATE  (--clean)
# ══════════════════════════════════════════════════════════════════════════════
CLEAN_MAP = [
    ("action",    "action"),   ("sum",     "sum"),     ("pm",       "pm"),
    ("pmplan",    "pmplan"),   ("abbr_old","abbr_old"),("abbr_new", "abbr_new"),
    ("roster",    "roster"),
]


def clean_template(src, dst):
    """สร้าง template เปล่า: ล้างข้อมูลเก่าทั้งหมด เก็บหัวตาราง + 1 แถวสไตล์ไว้ stamp"""
    shutil.copy2(src, dst)
    wb = load_workbook(dst)

    for lay_key, sheet_key in CLEAN_MAP:
        L = LAYOUT[lay_key]
        ws = wb[LAYOUT["sheets"][sheet_key]]
        start, (c0, c1) = L["row_start"], L["cols"]
        c1 = min(c1, ws.max_column)
        unmerge_in(ws, start + 1, ws.max_row)
        clear_block(ws, start, start, c0, c1)                        # แถวสไตล์: ล้างค่า เก็บฟอร์แมต
        clear_block(ws, start + 1, ws.max_row, c0, c1, reset_style=True)
        print(f"  🧹 {ws.title:<22} ล้างแถว {start}–{ws.max_row}")

    # Tools / Spares / Outsource
    Li = LAYOUT["itemlist"]
    for key, title in [("tools", "รายการเครื่องมือช่าง (Tools)"),
                       ("spares", "รายการอะไหล่และวัสดุสิ้นเปลือง (Spare Parts)")]:
        ws = wb[LAYOUT["sheets"][key]]
        w(ws, 1, 1, title)
        clear_block(ws, Li["row_start"], Li["row_start"], *Li["cols"])
        clear_block(ws, Li["row_start"] + 1, ws.max_row, *Li["cols"], reset_style=True)
        print(f"  🧹 {ws.title:<22} ล้างแถว {Li['row_start']}–{ws.max_row}")

    Lo = LAYOUT["outsource"]
    ws = wb[LAYOUT["sheets"]["outsource"]]
    clear_block(ws, Lo["row_start"], Lo["row_start"], *Lo["cols"])
    clear_block(ws, Lo["row_start"] + 1, ws.max_row, *Lo["cols"], reset_style=True)
    print(f"  🧹 {ws.title:<22} ล้างแถว {Lo['row_start']}–{ws.max_row}")

    wb.save(dst)
    print(f"\n💾 template เปล่า: {dst}")
    return dst


# ══════════════════════════════════════════════════════════════════════════════
#  SECTION 7 — DIAGNOSTICS + MAIN
# ══════════════════════════════════════════════════════════════════════════════
def dump_layout(path, sheet, max_row=12, max_col=60):
    wb = load_workbook(path)
    ws = wb[sheet]
    print(f"### SHEET: {sheet}  max_row={ws.max_row} max_col={ws.max_column}")
    print("### MERGED (30):", [str(x) for x in list(ws.merged_cells.ranges)[:30]])
    for row in ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col):
        for c in row:
            if c.value is not None:
                v = str(c.value)
                print(f"{c.coordinate} | {v[:70]}")


def dump_all_sheets(path):
    wb = load_workbook(path)
    print(f"### {os.path.basename(path)}")
    for s in wb.sheetnames:
        ws = wb[s]
        print(f"  {s!r:28} rows={ws.max_row:<5} cols={ws.max_column:<4} merged={len(ws.merged_cells.ranges)}")


def save_to_drive(local_path, folder="AMR_EstCost"):
    try:
        from google.colab import drive
        drive.mount("/content/drive", force_remount=False)
        dest_dir = f"/content/drive/My Drive/{folder}"
        os.makedirs(dest_dir, exist_ok=True)
        dest = os.path.join(dest_dir, os.path.basename(local_path))
        shutil.copy2(local_path, dest)
        print(f"📁 Drive: My Drive/{folder}/{os.path.basename(local_path)}")
        return dest
    except Exception as e:
        print(f"⚠️  Drive: {e}")
        return None


def main(template_path, input_path, output_dir="."):
    print(f"\n📖 อ่าน input: {input_path}")
    cfg = load_input(input_path)

    err, warn = validate(cfg)
    for x in warn:
        print(f"  ⚠️  {x}")
    for x in err:
        print(f"  ❌ {x}")
    if err:
        print("\n⛔ หยุด — แก้ input ก่อน")
        return None

    print("🧮 คำนวณชั่วโมง + กระจาย 12 เดือน...")
    rows = calc_equipment_hours(cfg)
    month_totals = balance_months(rows)

    # ชื่อไฟล์ผลลัพธ์: project_short_name > project_name > ชื่อไฟล์ input
    stem = (cfg["project"].get("project_short_name") or "").strip()
    if not stem:
        stem = cfg["project"]["project_name"].strip()[:60]
    if not stem:
        stem = re.sub(r"^input[_ ]?data[_ ]?", "",
                      os.path.splitext(os.path.basename(input_path))[0], flags=re.I)
    stem = re.sub(r'[\\/*?:"<>|\r\n\t]', "", stem).strip() or "output"
    out = os.path.join(output_dir, f"Est_Cost_MA_{stem}.xlsx")
    print(f"📋 Copy template → {out}")
    os.makedirs(output_dir, exist_ok=True)   # สร้างโฟลเดอร์ปลายทางถ้ายังไม่มี
    shutil.copy2(template_path, out)

    wb = load_workbook(out)
    print("🔧 patching ทุก sheet...")
    patch_summary(wb, cfg)
    patch_tools(wb, cfg)
    patch_spares(wb, cfg)
    patch_outsource(wb, cfg)
    patch_telephone(wb, cfg)
    patch_mprate(wb, cfg)
    action_map = patch_action(wb, cfg, rows)
    sum_rowmap = patch_sum(wb, cfg, rows, action_map)
    patch_pm_schedule(wb, cfg, rows, sum_rowmap, month_totals)
    patch_pm_plan(wb, cfg, rows)
    patch_abbreviation(wb, cfg, rows)
    patch_roster(wb, cfg)
    patch_shift_plan(wb, cfg, rows)
    patch_routing_images(wb, cfg, output_dir)
    patch_pm_routing(wb, cfg, rows, svg_dir=os.path.join(output_dir, f'route_svg_{stem}'))

    wb.save(out)
    print(f"\n✅ บันทึก: {out}")
    print(f"   ชั่วโมง PM รวมทั้งปี: {sum(r['annual_hours'] for r in rows):,.0f} MH")
    save_to_drive(out)
    return out


if __name__ == "__main__":
    a = sys.argv[1:]
    if not a:
        print(__doc__)
    elif a[0] == "--check":
        cfg = load_input(a[1])
        e, wn = validate(cfg)
        for x in wn: print(f"⚠️  {x}")
        for x in e:  print(f"❌ {x}")
        if not e and not wn: print("✅ input ผ่านทุกข้อ")
    elif a[0] == "--clean":
        clean_template(a[1], a[2] if len(a) > 2 else "Est_Cost_MA_v5_clean.xlsx")
    elif a[0] == "--dump":
        dump_layout(a[1], a[2]) if len(a) >= 3 else dump_all_sheets(a[1])
    elif len(a) >= 2:
        main(a[0], a[1], a[2] if len(a) > 2 else ".")
    else:
        print("Usage: python generate_v5.py <template.xlsx> <input.xlsx>")
